"""
    项目中部门管理相关业务功能用到的views视图函数
"""

from django.shortcuts import render, redirect, HttpResponse
from django.http import JsonResponse
from app_web import models
from django import forms
from django.core.validators import RegexValidator
from django.core.exceptions import ValidationError
from django.utils.safestring import mark_safe
from app_web.utils.pagination import Pagination
from app_web.utils.bootstrap_mf import BootStrapModelForm
from app_web.utils.model_form import EmployeeModelForm, PrettyNumModelForm, PrettyNumEditModelForm, DepartmentModelForm


def department_list(request):
    """ 部门列表 """
    # # 检查用户是否已经登录,已登录再继续执行代码,未登录的话跳转到登录页面.
    # login_info = request.session.get("login_info")   # 获取存储登录信息的字典里面的信息
    # if not login_info:
    #     return redirect('/login/')

    # 去数据库中获取所有的部门列表
    # queryset类型是:列表中放多个对象,每个对象中封装着他的一行的数据    [对象,对象]
    queryset_a = models.Department.objects.all()

    paginate_object_a = Pagination(request, queryset_a, page_size=5)

    page_context = {"queryset_a": paginate_object_a.paginate_queryset,
                    "page_string": paginate_object_a.html_pagination(),
                    }

    return render(request, 'department_list.html', page_context)
    # 把queryset_a当做参数传递给render,django在内部模版渲染的时候就会取到queryset_a


def department_add(request):
    """ 添加部门 """
    if request.method == "GET":
        return render(request, 'department_add.html')

    # 获取用户通过POST提交过来的数据(暂时忽略title内容为空)
    submit_title = request.POST.get("title")

    # 保存到数据库
    models.Department.objects.create(title=submit_title)

    # 重定向回到部门列表页面
    return redirect("/department/list/")
    # 测试 :return render(request, 'department_list.html')


def department_model_form_add(request):
    """ 新建部门 (ModelForm版本)"""
    if request.method == "GET":
        form_a = DepartmentModelForm()
        return render(request, 'department_model_form_add.html', {"form": form_a})

    # 用户POST提交数据,数据校验.
    form_a = DepartmentModelForm(data=request.POST)
    if form_a.is_valid():
        # 如果数据合法,然后保存到数据库
        # {'employee_name': 'jason', 'employee_password': '12366', 'employee_age': 23, 'employee_account': Decimal('1200'), 'create_time': datetime.datetime(2016, 8, 15, 0, 0, tzinfo=zoneinfo.ZoneInfo(key='UTC')), 'gender': 1, 'department': <Department: IT运维部>}
        # print(form_a.cleaned_data)
        form_a.save()  # form_a会将数据存储到上面定义的model这个类里面
        return redirect('/department/list/')
    # 校验失败(在页面上显示错误信息)
    return render(request, 'department_model_form_add.html', {"form": form_a})


def department_delete(request):
    """删除部门"""

    # 新加功能:在点击删除按钮后,要弹出确认是否要删除的提示框.
    # 获取到要删除管理员的ID
    department_id = request.GET.get('department_id')

    # 校验要删除的管理员是否存在?
    department_exists = models.Department.objects.filter(id=department_id).exists()
    # 如果在库中查询到不存在,就返回错误信息
    if not department_exists:
        return JsonResponse({"status": False, 'error': "部门数据不存在, 删除失败"})
    # 管理员数据存在就到数据库中进行删除
    models.Department.objects.filter(id=department_id).delete()
    return JsonResponse({"status": True})


    # 删除ID   http://127.0.0.1:8000/department/delete/?nid=1
    # nid = request.GET.get('nid')
    # 删除
    # models.Department.objects.filter(id=nid).delete()
    # 重定向回到部门列表
    # return redirect("/department/list/")


def department_edit(request, nid):
    """ 修改部门 """

    if request.method == "GET":
        # 根据nid获取当前部门的数据信息
        row_object = models.Department.objects.filter(id=nid).first()
        # print(row_object.id, row_object.title)
        return render(request, 'department_edit.html', {"row_object": row_object})

    # 获取用户提交的标题
    post_title = request.POST.get("title")

    # 根据ID找到数据库中的数据并进行更新.
    models.Department.objects.filter(id=nid).update(title=post_title)

    # 重定向回到部门列表
    return redirect("/department/list/")


def department_batch_upload(request):
    """ 部门列表 批量上传函数 (基于Excel文件)"""

    from openpyxl import load_workbook

    # 1. 获取用户上传的文件对象
    file_object = request.FILES.get("upload_file")
    # print(type(file_object))  # 后台打印 <class 'django.core.files.uploadedfile.InMemoryUploadedFile'>

    # 2. 对象传递给openpyxl,它来读取文件的内容
    workbook_a = load_workbook(file_object)
    sheet_a = workbook_a.worksheets[0]

    # 3. 循环获取每一行数据
    for row_a in sheet_a.iter_rows(min_row=2):  # min_row=2表示从第二行开始.
        text_a = row_a[0].value
        # print(text_a)

        # 写入数据库
        # 先判断在数据库中是否存在
        exists_a = models.Department.objects.filter(title=text_a).exists()
        if not exists_a:  # 数据库中不存在,然后添加
            models.Department.objects.create(title=text_a)

    return redirect("/department/list/")





